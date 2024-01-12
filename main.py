from youtube_link import youtube_links,description
from spotify import get_token
from spotify import search_spotify
from Get_Ghazals import get_ghazal
from singer import singername
from youtube_link_from_rekhta import get_youtube_link
import openpyxl
from tqdm import tqdm

def initiate():
    poet_name = input("Artist Name: ")

    data2=[]
    data_dump = []
    data2.append(("Title","Links","Singer"))
    data_dump.append(("Title", "Links", "Singer", "View Count"))

    ghazals = get_ghazal(poet_name)
    token = get_token()

    api_key_list = []
    api_index = 0
    with open ("api_keys.txt", "r") as f:
        for line in f:
            api_key_list.append(line.strip("\n"))

    try:
        for ghazal in tqdm(ghazals):
            ghazals[ghazal]["Youtube Links"] = get_youtube_link(ghazals[ghazal]["Rekhta Link"])
            count = 0
            for i in ghazals[ghazal]["Youtube Links"]:
                if not count:
                    det = (ghazal, i, ghazals[ghazal]["Youtube Links"][i])
                else:
                    det = ("", i, ghazals[ghazal]["Youtube Links"][i])
                count += 1
                data2.append(det)
            if count == 0:
                data2.append((ghazal, "", ""))
            ghazals[ghazal]["Youtube Links"] = ["Quote Exceeded"]
            ghazals[ghazal]["Youtube Links"] = youtube_links(poet_name, ghazal, api_key_list[api_index])
            while ghazals[ghazal]["Youtube Links"] == ["Quote Exceeded"]:
                api_index += 1
                ghazals[ghazal]["Youtube Links"] = youtube_links(poet_name, ghazal, api_key_list[api_index])
            if '' in ghazals[ghazal]["Youtube Links"]:
                ghazals[ghazal]["Youtube Links"].remove('')
            ghazals[ghazal]["Spotify Links"] = search_spotify(token, ghazal)
            yt_links = {}
            vals = []
            for i in range(len(ghazals[ghazal]["Youtube Links"])):
                desc = description(ghazals[ghazal]["Youtube Links"][i], api_key_list[api_index])
                while desc == "Quote Exceeded":
                    api_index += 1
                    desc = description(ghazals[ghazal]["Youtube Links"][i], api_key_list[api_index])
                singer = singername(desc[0])
                vals.append(desc[1])
                if singer != "Unknown":
                    if singer not in yt_links:
                        yt_links[singer] = [ghazals[ghazal]["Youtube Links"][i], desc[1]]
                    elif yt_links[singer][1] < desc[1]:
                        yt_links[singer] = [ghazals[ghazal]["Youtube Links"][i], desc[1]]
                if i == 0:
                    data_dump.append((ghazal, ghazals[ghazal]["Youtube Links"][i], singer, desc[1]))
                else:
                    data_dump.append(("", ghazals[ghazal]["Youtube Links"][i], singer, desc[1]))
            if yt_links:
                singers = list(yt_links.keys())
                values = list(yt_links.values())
            else:
                singer = "Unknown"
            for i in range(len(ghazals[ghazal]["Spotify Links"][0])):
                if count == 0:
                    data_dump.append((ghazal, ghazals[ghazal]["Spotify Links"][0][i], ghazals[ghazal]["Spotify Links"][1][i], ""))
                else:
                    data_dump.append(("", ghazals[ghazal]["Spotify Links"][0][i], ghazals[ghazal]["Spotify Links"][1][i], ""))
                count += 1
            if len(ghazals[ghazal]["Youtube Links"]) == 0 and len(ghazals[ghazal]["Spotify Links"][0]) == 0:
                data_dump.append((ghazal, "", "", ""))
    except IndexError:
        data2.pop()
        print("Reached Youtube api limit so unable to load complete data and writing partial data for the writer in the excel file.")
    generate_file(poet_name, ghazals, data_dump, data2)

def generate_file(poet_name, ghazals, data_dump, data2):
    wb = openpyxl.Workbook()
    sheet3=wb.active
    sheet3.title = "Youtube and Spotify"
    row = 1
    for i in tqdm(data_dump):
        if i[0] != "" and i[0] != "Title":
            for j in range(1, 5):
                cur = sheet3.cell(row=row, column=j)
                if j == 1:
                    cur.value = i[0]; cur.hyperlink = ghazals[i[0]]["Rekhta Link"]
                elif j == 2 and i[1] != "":
                    cur.value = i[1]; cur.hyperlink = i[1]
                else:
                    cur.value = i[j - 1]
        elif i[1] != "" and i[1] != "Links":
            for j in range(1, 5):
                cur = sheet3.cell(row=row, column=j)
                if j == 2:
                    cur.value = i[1]; cur.hyperlink = i[1]
                else:
                    cur.value = i[j - 1]
        else:
            sheet3.append(i)
        row += 1

    row=1
    sheet2=wb.create_sheet("Rekhta")
    for i in tqdm(data2):
        if i[0]!="" and i[0]!="Title":
            for j in range(1,4):
                cur=sheet2.cell(row=row,column=j)
                if j==1:
                    cur.value=i[0];cur.hyperlink=ghazals[i[0]]["Rekhta Link"]
                elif j == 2 and i[1] != "":
                    cur.value = i[1]; cur.hyperlink = i[1]
                else:
                    cur.value=i[j-1]
        elif i[1] != "" and i[1] != "Links":
            for j in range(1, 4):
                cur = sheet2.cell(row=row, column=j)
                if j == 2:
                    cur.value = i[1]; cur.hyperlink = i[1]
                else:
                    cur.value = i[j - 1]
        else:
            sheet2.append(i)
        row+=1

    excel_file = f"Poets\\{poet_name}.xlsx"
    wb.save(excel_file)

if __name__ == "__main__":
    initiate()